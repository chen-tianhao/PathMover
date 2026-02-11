#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Arrow Range:
    ("R1", "A12", "CO60"),
    ("R2", "A105", "CO146"),
    ("R3", "A193", "CO239"),
    ("R4", "AIE12", "AJE239"),
"""

import json
from collections import defaultdict
from typing import List, Tuple, Optional, Dict, Set
import pathlib
import re

import matplotlib.pyplot as plt
from matplotlib.patches import FancyArrowPatch
from matplotlib.lines import Line2D
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


BASE_DIR = pathlib.Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "F3 Traffic Config File.xlsx"
SHEET_NAME = "Traffic Config"
OUT_JSON = BASE_DIR / "control_points_v16.json"


def color_id(cell) -> Optional[str]:
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return None
    c = fill.fgColor
    if c is None:
        return None
    if c.type == "rgb" and c.rgb:
        return c.rgb
    if c.type == "indexed":
        return f"indexed:{c.indexed}"
    if c.type == "theme":
        return f"theme:{c.theme}:{c.tint}"
    return None


def contiguous_segments_for_color(ws, row: int, color_sig: str, col_min: int, col_max: int) -> List[Tuple[int, int]]:
    segs = []
    in_seg = False
    s = None
    for c in range(col_min, col_max + 1):
        if color_id(ws.cell(row=row, column=c)) == color_sig:
            if not in_seg:
                in_seg = True
                s = c
        else:
            if in_seg:
                segs.append((s, c - 1))
                in_seg = False
    if in_seg:
        segs.append((s, col_max))
    return segs


def sample_positions_including_end(x_start: float, x_end: float, step: float) -> List[float]:
    if x_end < x_start:
        x_start, x_end = x_end, x_start
    xs = [float(x_start)]
    x = x_start
    while x + step <= x_end + 1e-9:
        x += step
        xs.append(float(x))
    if abs(xs[-1] - x_end) > 1e-9:
        xs.append(float(x_end))
    return xs


def sample_int_positions_including_end(start: int, end: int, step: int) -> List[int]:
    if end < start:
        start, end = end, start
    ys = [start]
    y = start
    while y + step <= end:
        y += step
        ys.append(y)
    if ys[-1] != end:
        ys.append(end)
    return ys


def add_edge(next_map: Dict[str, Set[str]], a: str, b: Optional[str]):
    if b is None:
        return
    next_map.setdefault(a, set()).add(b)


def build_row_index(points: Dict[str, dict], region: str) -> Dict[int, Dict[float, str]]:
    out: Dict[int, Dict[float, str]] = defaultdict(dict)
    for pid, p in points.items():
        if p["region"] != region:
            continue
        if p.get("meta", {}).get("kind") not in ("h", "seg"):
            continue
        r = int(round(p["y"]))
        out[r][p["x"]] = pid
    return out


def build_rowseg_index(points: Dict[str, dict], region: str) -> Dict[Tuple[int, int], List[str]]:
    tmp: Dict[Tuple[int, int], List[Tuple[float, str]]] = defaultdict(list)
    for pid, p in points.items():
        if p["region"] != region:
            continue
        meta = p.get("meta", {})
        if meta.get("kind") != "seg":
            continue
        key = (int(round(p["y"])), int(meta["seg"]))
        tmp[key].append((p["x"], pid))
    out: Dict[Tuple[int, int], List[str]] = {}
    for key, lst in tmp.items():
        out[key] = [pid for _, pid in sorted(lst, key=lambda t: t[0])]
    return out


def build_col_index(points: Dict[str, dict], region: str) -> Dict[int, Dict[int, str]]:
    out: Dict[int, Dict[int, str]] = defaultdict(dict)
    for pid, p in points.items():
        if p["region"] != region:
            continue
        if p.get("meta", {}).get("kind") != "v":
            continue
        x = int(round(p["x"]))
        y = int(round(p["y"]))
        out[x][y] = pid
    return out


def ordered_in_row(points_by_row: Dict[int, Dict[float, str]], row: int) -> List[Tuple[float, str]]:
    d = points_by_row.get(row, {})
    return sorted(((x, pid) for x, pid in d.items()), key=lambda t: t[0])


def ordered_in_col(points_by_col: Dict[int, Dict[int, str]], x: int) -> List[Tuple[int, str]]:
    d = points_by_col.get(x, {})
    return sorted(((y, pid) for y, pid in d.items()), key=lambda t: t[0])


def cell_to_xy(cell_ref: str) -> Tuple[float, float]:
    """Excel cell like 'E13' -> (x, y) where x=col-1, y=row."""
    m = re.match(r"^([A-Z]+)(\d+)$", cell_ref.strip().upper())
    if not m:
        raise ValueError(f"Bad cell ref: {cell_ref}")
    col_letters, row_s = m.group(1), m.group(2)
    x = column_index_from_string(col_letters) - 1
    y = int(row_s)
    return float(x), float(y)


def in_box(p: dict, x_min: float, y_min: float, x_max: float, y_max: float) -> bool:
    return (x_min <= p["x"] <= x_max) and (y_min <= p["y"] <= y_max)


def main():
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    col_AIS = column_index_from_string("AIS")
    col_max_use = col_AIS - 1  # left of AIS is scanned for colored horizontal regions

    # Anchor colors
    purple_sig = color_id(ws["E13"])
    blue_sig = color_id(ws["A28"])
    orange_sig = color_id(ws["E28"])
    green_sig = color_id(ws["E120"])
    if any(s is None for s in [purple_sig, blue_sig, orange_sig, green_sig]):
        raise RuntimeError("Failed to detect one or more anchor colors (E13/A28/E28/E120).")

    points: Dict[str, dict] = {}
    coord_region: Dict[Tuple[float, float], str] = {}

    # -------- Generate points --------
    # ---------------- BLUE FIRST ----------------
    # Blue cells left of AIS
    for r in range(1, ws.max_row + 1):
        for c in range(1, col_max_use + 1):
            if color_id(ws.cell(row=r, column=c)) == blue_sig:
                x = float(c - 1)
                y = float(r)
                pid = f"BL_R{r:03d}_C{c:04d}"
                points[pid] = {"id": pid, "x": x, "y": y, "region": "blue",
                               "meta": {"kind": "cell", "row": r, "col": c}}
                coord_region[(x, y)] = "blue"

    # Blue cells in AIS–AIV
    col_AIV = column_index_from_string("AIV")
    for r in range(1, ws.max_row + 1):
        for c in range(col_AIS, col_AIV + 1):
            if color_id(ws.cell(row=r, column=c)) == blue_sig:
                x = float(c - 1)
                y = float(r)
                pid = f"BL_R{r:03d}_C{c:04d}"
                points[pid] = {"id": pid, "x": x, "y": y, "region": "blue",
                               "meta": {"kind": "cell", "row": r, "col": c}}
                coord_region[(x, y)] = "blue"
                
                
    # Horizontal purple (step=5)
    purple_rows = list(range(13, 17)) + list(range(22, 26)) + list(range(226, 230)) + list(range(235, 239))
    x_start = column_index_from_string("E") - 1
    x_end = column_index_from_string("AIL") - 1
    step_purple = 5.0
    for r in purple_rows:
        if color_id(ws.cell(row=r, column=column_index_from_string("E"))) != purple_sig:
            continue
        xs = sample_positions_including_end(x_start, x_end, step_purple)
        for i, x in enumerate(xs):
            y = float(r)
            key = (x, y)
            if key in coord_region and coord_region[key] == "blue":
                continue
            pid = f"PUH_R{r:03d}_{i:04d}"
            points[pid] = {"id": pid, "x": float(x), "y": float(r), "region": "purple", "meta": {"kind": "h", "row": r}}
            coord_region[key] = "purple"
            

    # Grey horizontal bands (step=2)
    step_grey = 2.0
    grey_bands = [{"row_min": 17, "row_max": 21, "y": 19}, {"row_min": 230, "row_max": 234, "y": 232}]
    for band in grey_bands:
        max_x = None
        for mr in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = mr.bounds
            if min_row == band["row_min"] and max_row == band["row_max"]:
                if min_col >= col_AIS:
                    continue
                x2 = max_col - 1
                max_x = x2 if max_x is None else max(max_x, x2)
        if max_x is not None:
            xs = sample_positions_including_end(4.0, float(max_x), step_grey)
            for i, x in enumerate(xs):
                y = float(band["y"])
                key = (float(x), y)
                # blue has highest priority
                if key in coord_region and coord_region[key] == "blue":
                    continue
                pid = f"GYH_Y{band['y']}_{i:04d}"
                points[pid] = {"id": pid, "x": float(x), "y": y, "region": "grey",
                               "meta": {"kind": "h", "band_y": band["y"]}}
                coord_region[key] = "grey"


    # Orange segments (step=5)
    step_orange = 5.0
    for r in range(1, ws.max_row + 1):
        segs = contiguous_segments_for_color(ws, r, orange_sig, 1, min(ws.max_column, col_max_use))
        for s_idx, (c1, c2) in enumerate(segs, start=1):
            xs = sample_positions_including_end(c1 - 1, c2 - 1, step_orange)
            for i, x in enumerate(xs):   
                y = float(r)
                key = (x, y)
                if key in coord_region and coord_region[key] == "blue":
                    continue
                pid = f"OR_R{r:03d}_S{s_idx:02d}_{i:04d}"
                points[pid] = {"id": pid, "x": float(x), "y": float(r), "region": "orange",
                               "meta": {"kind": "seg", "row": r, "seg": s_idx, "c1": c1, "c2": c2}}
                coord_region[key] = "orange"


    # Green segments (step=5)
    step_green = 5.0
    for r in range(1, ws.max_row + 1):
        segs = contiguous_segments_for_color(ws, r, green_sig, 1, min(ws.max_column, col_max_use))
        for s_idx, (c1, c2) in enumerate(segs, start=1):
            xs = sample_positions_including_end(c1 - 1, c2 - 1, step_green)
            for i, x in enumerate(xs):
                y = float(r)
                key = (x, y)
                if key in coord_region and coord_region[key] == "blue":
                    continue
                pid = f"GN_R{r:03d}_S{s_idx:02d}_{i:04d}"
                points[pid] = {"id": pid, "x": float(x), "y": float(r), "region": "green",
                               "meta": {"kind": "seg", "row": r, "seg": s_idx, "c1": c1, "c2": c2}}
                coord_region[key] = "green"
                
                
    # Vertical AIS-AJE explicit purple/grey
    step_v_purple = 5
    step_v_grey = 2
    vertical_ranges = [(28, 101), (150, 223)]

    def add_vertical_points(col_from: str, col_to: str, tag: str, region: str, step_v: int):
        col1 = column_index_from_string(col_from)
        col2 = column_index_from_string(col_to)
        for col in range(col1, col2 + 1):
            x = col - 1
            for (r1, r2) in vertical_ranges:
                ys = sample_int_positions_including_end(r1, r2, step_v)
                for i, y in enumerate(ys):
                    key = (x, y)
                    if key in coord_region and coord_region[key] == "blue":
                        continue
                    pid = f"{tag}_X{x:04d}_R{r1:03d}_{i:03d}"
                    points[pid] = {"id": pid, "x": float(x), "y": float(y), "region": region,
                                   "meta": {"kind": "v", "col": col, "r1": r1, "r2": r2}}
                    coord_region[key] = region
                    

    add_vertical_points("AIS", "AIT", "PUVUP", "purple", step_v_purple)
    add_vertical_points("AIU", "AIV", "PUVDN", "purple", step_v_purple)
    add_vertical_points("AJB", "AJE", "PUVDN", "purple", step_v_purple)
    add_vertical_points("AIY", "AIY", "GYV", "grey", step_v_grey)

    # -------- Connectivity --------
    next_map: Dict[str, Set[str]] = {}

    purple_row_index = build_row_index(points, "purple")
    orange_rowseg = build_rowseg_index(points, "orange")
    green_rowseg = build_rowseg_index(points, "green")
    purple_col_index = build_col_index(points, "purple")

    # Purple forward edges (horizontal)
    def connect_purple_rows(rows: List[int], direction: str):
        for r in rows:
            ordered = ordered_in_row(purple_row_index, r)
            if len(ordered) < 2:
                continue
            if direction == "right":
                for i in range(len(ordered)-1):
                    add_edge(next_map, ordered[i][1], ordered[i+1][1])
            else:
                for i in range(1, len(ordered)):
                    add_edge(next_map, ordered[i][1], ordered[i-1][1])

    connect_purple_rows(list(range(13, 17)) + list(range(22, 24)), "right")
    connect_purple_rows(list(range(24, 26)), "left")
    connect_purple_rows(list(range(226, 228)), "right")
    connect_purple_rows(list(range(228, 230)) + list(range(235, 239)), "left")

    # Purple lane-change helper (same x)
    def lane_change(r_from: int, r_to: int, allow: bool):
        if not allow:
            return
        a = purple_row_index.get(r_from, {})
        b = purple_row_index.get(r_to, {})
        if not a or not b:
            return
        for x, pid in a.items():
            pid2 = b.get(x)
            if pid2:
                add_edge(next_map, pid, pid2)

    # rows 13-16: up/down within band
    for r in range(13, 16):
        lane_change(r, r+1, True)
        lane_change(r+1, r, True)

    # rows 22-23: both ways; 23->24 forbidden
    lane_change(22, 23, True)
    lane_change(23, 22, True)
    # connect purple 23->24 both ways
    lane_change(23, 24, True)
    lane_change(24, 23, True)
    
    # rows 24-25: row24 down; row25 up
    lane_change(24, 25, True)
    lane_change(25, 24, True)
    
    # rows 226-227: row226 down; row227 up
    lane_change(226, 227, True)
    lane_change(227, 226, True)
    
    # rows 228-229: up/down; 228->227 forbidden (not added)
    lane_change(228, 229, True)
    lane_change(229, 228, True)
    # connect purple 228->227 up/down
    lane_change(227, 228, True)
    lane_change(228, 227, True)
    
    # rows 235-238: up/down within band
    for r in range(235, 238):
        lane_change(r, r+1, True)
        lane_change(r+1, r, True)

    # Orange/Green forward by row rule
    def og_dir(row: int) -> str:
        if row == 125:
            return "right"
        if row >= 126:
            return "left"
        return "right"

    def connect_segments(rowseg: Dict[Tuple[int, int], List[str]], direction_func):
        for (r, seg), ids in rowseg.items():
            if len(ids) < 2:
                continue
            d = direction_func(r)
            if d == "right":
                for i in range(len(ids)-1):
                    add_edge(next_map, ids[i], ids[i+1])
            else:
                for i in range(1, len(ids)):
                    add_edge(next_map, ids[i], ids[i-1])

    connect_segments(orange_rowseg, og_dir)
    connect_segments(green_rowseg, og_dir)

    # Orange/Green lane-change (same x) with boundary constraints
    def build_row_x_index(region: str) -> Dict[int, Dict[float, str]]:
        out = defaultdict(dict)
        for pid, p in points.items():
            if p["region"] != region:
                continue
            if p.get("meta", {}).get("kind") != "seg":
                continue
            out[int(round(p["y"]))][p["x"]] = pid
        return out

    orange_row_x = build_row_x_index("orange")
    green_row_x = build_row_x_index("green")

    def lane_change_og(row_x: Dict[int, Dict[float, str]], r_from: int, r_to: int, allow: bool):
        if not allow:
            return
        a = row_x.get(r_from, {})
        b = row_x.get(r_to, {})
        if not a or not b:
            return
        for x, pid in a.items():
            pid2 = b.get(x)
            if pid2:
                add_edge(next_map, pid, pid2)

    all_og_rows = sorted(set(orange_row_x.keys()) | set(green_row_x.keys()))
    
    for r in all_og_rows:
        # connect adjacent rows
        candidates = [r - 1, r + 1]
    
        for r_to in candidates:
            if r_to not in all_og_rows:
                continue
    
            # row 125 cannot reach 126, vice versa
            if (r == 125 and r_to == 126) or (r == 126 and r_to == 125):
                allow = False
            else:
                allow = True
    
            lane_change_og(orange_row_x, r, r_to, allow)
            lane_change_og(green_row_x,  r, r_to, allow)

    # -------- BLUE connectivity --------
    # Build a coordinate index. NOTE: (x,y) may contain BOTH blue and purple points after Excel updates.
    coord_index: Dict[Tuple[float, float], List[str]] = defaultdict(list)
    for pid, p in points.items():
        coord_index[(p["x"], p["y"])].append(pid)

    def get_pid_at(x: float, y: float, prefer_region: Optional[str] = None, exclude_region: Optional[str] = None) -> Optional[str]:
        cands = coord_index.get((x, y), [])
        if not cands:
            return None
        if exclude_region is not None:
            cands = [pid for pid in cands if points[pid]["region"] != exclude_region]
            if not cands:
                return None
        if prefer_region is not None:
            for pid in cands:
                if points[pid]["region"] == prefer_region:
                    return pid
        return cands[0]

    # Determine left/right direction for a given row, consistent with outside lanes
    right_purple_rows = set(list(range(13, 17)) + list(range(22, 24)) + list(range(226, 228)))
    left_purple_rows = set(list(range(24, 26)) + list(range(228, 230)) + list(range(235, 239)))

    def row_lr_dir(r: int) -> str:
        # purple-defined rows
        if r in right_purple_rows:
            return "right"
        if r in left_purple_rows:
            return "left"
        # orange/green rows (same as og_dir)
        return og_dir(r)  # "right" or "left"

    # 1) Horizontal links inside each BLUE segment + connect to outside point if exists
    # Use Excel scan to find BLUE segments in each row (contiguous)
    col_max_use_blue = min(ws.max_column, column_index_from_string("AIV"))
    for r in range(1, ws.max_row + 1):
        segs = contiguous_segments_for_color(ws, r, blue_sig, 1, col_max_use_blue)
        if not segs:
            continue

        d = row_lr_dir(r)  # "right" or "left"

        for (c1, c2) in segs:
            # Collect blue point ids in this segment (they are at every cell x=c-1)
            ids = []
            for c in range(c1, c2 + 1):
                x = float(c - 1)
                pid = get_pid_at(x, float(r), prefer_region="blue")
                if pid and points[pid]["region"] == "blue":
                    ids.append(pid)

            if len(ids) < 2:
                continue

            # Internal blue links
            # Sort by x
            ids_sorted = sorted(ids, key=lambda pid: points[pid]["x"])
            if d == "right":
                for i in range(len(ids_sorted) - 1):
                    add_edge(next_map, ids_sorted[i], ids_sorted[i + 1])
            else:
                for i in range(1, len(ids_sorted)):
                    add_edge(next_map, ids_sorted[i], ids_sorted[i - 1])

            # Connect to outside neighbor if exists
            x_left = float((c1 - 1) - 1)   # column (c1-1) => x=(c1-2)
            x_right = float((c2 + 1) - 1)  # column (c2+1) => x=c2
            outside_left = get_pid_at(x_left, float(r), exclude_region="blue")
            outside_right = get_pid_at(x_right, float(r), exclude_region="blue")

            if d == "right":
                # outside_left -> first_blue, last_blue -> outside_right
                if outside_left and points[outside_left]["region"] != "blue":
                    add_edge(next_map, outside_left, ids_sorted[0])
                if outside_right and points[outside_right]["region"] != "blue":
                    add_edge(next_map, ids_sorted[-1], outside_right)
            else:
                # outside_right -> first_blue (in right-to-left sense), last_blue -> outside_left
                if outside_right and points[outside_right]["region"] != "blue":
                    add_edge(next_map, outside_right, ids_sorted[-1])
                if outside_left and points[outside_left]["region"] != "blue":
                    add_edge(next_map, ids_sorted[0], outside_left)

    # 2) Vertical links in BLUE: each blue segment has 4 columns.
    #    First two columns go UP; last two go DOWN.
    #    Also connect across gaps between separate blue blocks (e.g., row31 <-> row54) by linking to nearest next blue point in direction.

    # Precompute BLUE segments per row for direction lookup
    blue_segs_by_row: Dict[int, List[Tuple[int, int]]] = {}
    for r in range(1, ws.max_row + 1):
        segs = contiguous_segments_for_color(ws, r, blue_sig, 1, col_max_use_blue)
        if segs:
            blue_segs_by_row[r] = segs

    def blue_vertical_dir(row: int, col: int) -> Optional[str]:
        """Return 'up' for first two cols of a 4-col segment, 'down' for last two, else None."""
        segs = blue_segs_by_row.get(row)
        if not segs:
            return None
        for (c1, c2) in segs:
            if c1 <= col <= c2:
                offset = col - c1
                # assume 4-column segments: offset 0,1 => up; 2,3 => down
                return "up" if offset <= 1 else "down"
        return None

    blue_points_by_x: Dict[int, List[Tuple[int, str, str]]] = defaultdict(list)  # x -> [(y,pid,dir)]
    for pid, p in points.items():
        if p["region"] != "blue":
            continue
        row = int(p["meta"]["row"])
        col = int(p["meta"]["col"])
        d = blue_vertical_dir(row, col)
        if d is None:
            continue
        x = int(round(p["x"]))
        y = int(round(p["y"]))
        blue_points_by_x[x].append((y, pid, d))

    # Explicitly connect BLUE points between columns AIP and AIS (same y), bidirectional.
    x_AIP = column_index_from_string("AIP") - 1
    x_AIS = column_index_from_string("AIS") - 1

    if x_AIP in blue_points_by_x and x_AIS in blue_points_by_x:
        aip_map = {y: pid for (y, pid, _) in blue_points_by_x[x_AIP]}
        ais_map = {y: pid for (y, pid, _) in blue_points_by_x[x_AIS]}
        for y in set(aip_map.keys()) & set(ais_map.keys()):
            r = int(y)
        
            d = row_lr_dir(r)  # "right" or "left"
        
            if d == "right":
                # AIP -> AIS
                add_edge(next_map, aip_map[y], ais_map[y])
            else:
                # AIS -> AIP
                add_edge(next_map, ais_map[y], aip_map[y])

    # Vertical purple forward edges (no cross-range links)
    x_AIS = column_index_from_string("AIS") - 1
    x_AIT = column_index_from_string("AIT") - 1
    x_AIU = column_index_from_string("AIU") - 1
    x_AIV = column_index_from_string("AIV") - 1
    x_AJB = column_index_from_string("AJB") - 1
    x_AJE = column_index_from_string("AJE") - 1

    for x, lst in blue_points_by_x.items():
        # Split into up/down columns and sort by y
        ups = sorted([(y, pid) for (y, pid, d) in lst if d == "up"], key=lambda t: t[0])
        dns = sorted([(y, pid) for (y, pid, d) in lst if d == "down"], key=lambda t: t[0])

        # UP columns: connect from larger y to next smaller y (nearest existing blue point)
        for i in range(1, len(ups)):
            y_prev, pid_prev = ups[i - 1]  # smaller y
            y_curr, pid_curr = ups[i]      # larger y
            add_edge(next_map, pid_curr, pid_prev)

        # DOWN columns: connect from smaller y to next larger y
        for i in range(len(dns) - 1):
            y_curr, pid_curr = dns[i]
            y_next, pid_next = dns[i + 1]
            add_edge(next_map, pid_curr, pid_next)


    for (r1, r2) in vertical_ranges:
        # AIS-AIT: UP
        for x in range(x_AIS, x_AIT + 1):
            ordered = [(y, pid) for (y, pid) in ordered_in_col(purple_col_index, x) if r1 <= y <= r2]
            for i in range(1, len(ordered)):
                if ordered[i][0] - ordered[i-1][0] <= step_v_purple + 1e-9:
                    add_edge(next_map, ordered[i][1], ordered[i-1][1])

        # AIU-AIV and AJB-AJE: DOWN
        for x in list(range(x_AIU, x_AIV + 1)) + list(range(x_AJB, x_AJE + 1)):
            ordered = [(y, pid) for (y, pid) in ordered_in_col(purple_col_index, x) if r1 <= y <= r2]
            for i in range(len(ordered) - 1):
                if ordered[i+1][0] - ordered[i][0] <= step_v_purple + 1e-9:
                    add_edge(next_map, ordered[i][1], ordered[i+1][1])

    # --- AIS-AIV: Directed adjacency links between BLUE cells and PURPLE vertical points (strict dy=1) ---
    # Only within columns AIS-AIV. Direction follows the PURPLE vertical direction:
    #   - AIS..AIT: UP (y decreases)  => edges go from y -> y-1
    #   - AIU..AIV: DOWN (y increases)=> edges go from y -> y+1
    #
    # This links ONLY immediately adjacent BLUE/PURPLE pairs in the same column (no diagonals, no nearest search),
    # and will not connect the two separated vertical segments because dy must be exactly 1.

    blue_index = {
        (int(round(float(p["x"]))), int(round(float(p["y"])))): pid
        for pid, p in points.items()
        if p["region"] == "blue" and p.get("meta", {}).get("kind") == "cell"
    }


    # --- NEW: connect BLUE vertical movement within AIS-AIV so upper/lower blue segments are connected ---
    # Blue in AIS-AIV follows the same directional split as the vertical corridors:
    #   - AIS..AIT columns: move UP (y -> y-1)
    #   - AIU..AIV columns: move DOWN (y -> y+1)
    x_AIS_v = column_index_from_string("AIS") - 1
    x_AIT_v = column_index_from_string("AIT") - 1
    x_AIU_v = column_index_from_string("AIU") - 1
    x_AIV_v = column_index_from_string("AIV") - 1

    for x in range(x_AIS_v, x_AIT_v + 1):  # UP columns
        ys = sorted({y for (xx, y) in blue_index.keys() if xx == x})
        for y in ys:
            pid = blue_index.get((x, y))
            # connect to nearest next BLUE above within 1..5 rows (bridges small gaps)
            for d in range(1, 6):
                pid2 = blue_index.get((x, y - d))
                if pid2:
                    add_edge(next_map, pid, pid2)
                    break

    for x in range(x_AIU_v, x_AIV_v + 1):  # DOWN columns
        ys = sorted({y for (xx, y) in blue_index.keys() if xx == x})
        for y in ys:
            pid = blue_index.get((x, y))
            # connect to nearest next BLUE below within 1..5 rows (bridges small gaps)
            for d in range(1, 6):
                pid2 = blue_index.get((x, y + d))
                if pid2:
                    add_edge(next_map, pid, pid2)
                    break


    # purple_pid -> (best_d, blue_pid)
    best_blue_to_purple = {}
    MAX_DY = 5  # 允许相邻距离 1~5
    
    def add_adj_bp_edge_up(xi: int, y_from: int):
        """UP direction: connect across (y_from -> y_from-d), choose nearest d in 1..MAX_DY."""
        # blue(y_from) -> purple(y_from-d)
        pid_b_from = blue_index.get((xi, y_from))
        if pid_b_from:
            for d in range(1, MAX_DY + 1):
                y_to = y_from - d
                if y_to < 0:
                    break
                pid_p_to = purple_col_index.get(xi, {}).get(y_to)
                if pid_p_to:
                    cur = best_blue_to_purple.get(pid_p_to)
                    # keep only the nearest blue (smaller d). tie-break: smaller |y_from| then blue id
                    cand = (d, abs(y_from), pid_b_from)
                    if cur is None or cand < cur[0]:
                        best_blue_to_purple[pid_p_to] = (cand, pid_b_from)
                    break
    
        # purple(y_from) -> blue(y_from-d)
        pid_p_from = purple_col_index.get(xi, {}).get(y_from)
        if pid_p_from:
            for d in range(1, MAX_DY + 1):
                y_to = y_from - d
                if y_to < 0:
                    break
                pid_b_to = blue_index.get((xi, y_to))
                if pid_b_to:
                    add_edge(next_map, pid_p_from, pid_b_to)   # 单向：沿UP
                    break
    
    
    def add_adj_bp_edge_down(xi: int, y_from: int):
        """DOWN direction: connect across (y_from -> y_from+d), choose nearest d in 1..MAX_DY."""
        # blue(y_from) -> purple(y_from+d)
        pid_b_from = blue_index.get((xi, y_from))
        if pid_b_from:
            for d in range(1, MAX_DY + 1):
                y_to = y_from + d
                pid_p_to = purple_col_index.get(xi, {}).get(y_to)
                if pid_p_to:
                    cur = best_blue_to_purple.get(pid_p_to)
                    cand = (d, abs(y_from), pid_b_from)
                    if cur is None or cand < cur[0]:
                        best_blue_to_purple[pid_p_to] = (cand, pid_b_from)
                    break
    
        # purple(y_from) -> blue(y_from+d)
        pid_p_from = purple_col_index.get(xi, {}).get(y_from)
        if pid_p_from:
            for d in range(1, MAX_DY + 1):
                y_to = y_from + d
                pid_b_to = blue_index.get((xi, y_to))
                if pid_b_to:
                    add_edge(next_map, pid_p_from, pid_b_to)   # 单向：沿DOWN
                    break


    # Iterate all y present in either BLUE or PURPLE on each column
    for x in range(x_AIS, x_AIT + 1):  # UP group
        ys = set(purple_col_index.get(x, {}).keys())
        ys |= {y for (xx, y) in blue_index.keys() if xx == x}
        for y in ys:
            add_adj_bp_edge_up(x, int(y))

    for x in range(x_AIU_v, x_AIV_v + 1):  # DOWN group
        ys = set(purple_col_index.get(x, {}).keys())
        ys |= {y for (xx, y) in blue_index.keys() if xx == x}
        for y in ys:
            add_adj_bp_edge_down(x, int(y))
    
    # finally commit BLUE -> PURPLE edges (each purple keeps only nearest blue)
    for pid_pu, (cand, pid_bl) in best_blue_to_purple.items():
        add_edge(next_map, pid_bl, pid_pu)
                    
        
    # Vertical lane-change between adjacent columns
    def lane_change_cols(x_from: int, x_to: int, allow: bool):
        if not allow:
            return
        for y, pid in purple_col_index.get(x_from, {}).items():
            pid2 = purple_col_index.get(x_to, {}).get(y)
            if pid2:
                add_edge(next_map, pid, pid2)

    # AIS can go right; AIT can go left
    lane_change_cols(x_AIS, x_AIS + 1, True)
    lane_change_cols(x_AIT, x_AIT - 1, True)

    # AIU-AIV: allow both sides
    for x in range(x_AIU_v, x_AIV_v + 1):
        if x == x_AIU:
            lane_change_cols(x, x + 1, True)
        else:
            lane_change_cols(x, x - 1, True)
            lane_change_cols(x, x + 1, True)
    
    # connect purple AIT & AIU
    lane_change_cols(x_AIT, x_AIU, True)  # AIT -> AIU
    lane_change_cols(x_AIU, x_AIT, True)  # AIU -> AIT
            
    # AJB-AJE: allow both sides
    for x in range(x_AJB, x_AJE + 1):
        if x == x_AJB:
            lane_change_cols(x, x + 1, True)
        elif x == x_AJE:
            lane_change_cols(x, x - 1, True)
        else:
            lane_change_cols(x, x - 1, True)
            lane_change_cols(x, x + 1, True)



    # -------- Right-side column AIV-AJB linking (reworked, no horizontal-grey-to-AIS bug) --------
    # Spec (latest):
    # 1) For ALL VERTICAL GREY points (kind='v'), find nearest LEFT boundary point on column AIV
    #    (must be purple OR blue) and nearest RIGHT boundary point on column AJB (must be purple),
    #    and link bidirectionally.
    # 2) Horizontal GREY bands (kind='h', rows 17-21 and 230-234) DO NOT connect to AIS/AJB columns.
    #    They connect ONLY to purple/blue on boundary rows (16 & 22) or (229 & 235) by nearest X.
    # 3) Force: row 22 blue cells (E..AIL) must connect to the nearest grey_h ABOVE (band 17-21) by nearest X.
    #           row 229 blue cells (E..AIL) must connect to the nearest grey_h BELOW (band 230-234) by nearest X.
    # 4) Then for BLUE points on column AIV that still have NO grey neighbor, find nearest GREY point(s)
    #    to its RIGHT (considering BOTH grey_v and grey_h). If 1 nearest -> link bidirectional.
    #    If 2 tie -> link to both bidirectional.

    import bisect

    EPS = 1e-9
    x_AIV = column_index_from_string("AIV") - 1
    x_AJB = column_index_from_string("AJB") - 1
    x_E   = column_index_from_string("E") - 1
    x_AIL = column_index_from_string("AIL") - 1

    # Collect boundary candidates on exact columns
    aiv_candidates = []  # [(pid, x, y)]
    ajb_candidates = []  # [(pid, x, y)]
    for pid, p in points.items():
        xi = int(round(p["x"]))
        if xi == x_AIV and p["region"] in ("purple", "blue"):
            yi = int(round(float(p["y"])))
            if p["region"] == "blue" and 102 <= yi <= 149:
                continue  # middle AIV blue does NOT connect to grey
            aiv_candidates.append((pid, float(p["x"]), float(p["y"])))
        if xi == x_AJB and p["region"] == "purple":
            ajb_candidates.append((pid, float(p["x"]), float(p["y"])))

    # Split grey points: vertical vs horizontal bands
    grey_v = []  # [(gid, x, y)]  kind='v' only
    grey_h = []  # [(gid, x, y)]  kind='h' in row-bands only
    for gid, gp in points.items():
        if gp["region"] != "grey":
            continue
        gx, gy = float(gp["x"]), float(gp["y"])
        kind = gp.get("meta", {}).get("kind")
        ry = int(round(gy))
        if kind == "v":
            grey_v.append((gid, gx, gy))
        elif kind == "h" and (17 <= ry <= 21 or 230 <= ry <= 234):
            grey_h.append((gid, gx, gy))

    all_grey = grey_v + grey_h

    # Helper: nearest by y (return 1 or 2 if tie), capped
    def nearest_by_y_ids(cands, y, cap=2):
        if not cands:
            return []
        c_sorted = sorted(cands, key=lambda t: t[2])
        ys = [t[2] for t in c_sorted]
        i = bisect.bisect_left(ys, y)
        neigh = []
        for j in (i - 1, i):
            if 0 <= j < len(c_sorted):
                neigh.append(c_sorted[j])
        if not neigh:
            return []
        best = None
        best_ids = []
        for pid, cx, cy in neigh:
            d = abs(cy - y)
            if best is None or d < best - EPS:
                best = d
                best_ids = [pid]
            elif abs(d - best) <= EPS:
                best_ids.append(pid)
        out = []
        for pid in best_ids:
            if pid not in out:
                out.append(pid)
        return out[:cap]

    # Helper: nearest by x (return 1 or 2 if tie), capped
    def nearest_by_x_ids(cands, x, cap=2):
        if not cands:
            return []
        best = None
        best_ids = []
        for pid, cx, cy in cands:
            d = abs(cx - x)
            if best is None or d < best - EPS:
                best = d
                best_ids = [pid]
            elif abs(d - best) <= EPS:
                best_ids.append(pid)
        out = []
        for pid in best_ids:
            if pid not in out:
                out.append(pid)
        return out[:cap]

    # --- Step 1: VERTICAL grey strip links to AIV/AJB by nearest Y (bidirectional) ---
    for gid, gx, gy in grey_v:
        for pid in nearest_by_y_ids(aiv_candidates, gy, cap=2):
            add_edge(next_map, gid, pid)
            add_edge(next_map, pid, gid)
        for pid in nearest_by_y_ids(ajb_candidates, gy, cap=2):
            add_edge(next_map, gid, pid)
            add_edge(next_map, pid, gid)

    # --- Horizontal grey bands link ONLY to boundary rows by nearest X (bidirectional) ---
    pb_row16, pb_row22, pb_row229, pb_row235 = [], [], [], []
    for pid, p in points.items():
        if p["region"] not in ("purple", "blue"):
            continue
        y = int(round(float(p["y"])))
        if y == 16:
            pb_row16.append((pid, float(p["x"]), float(p["y"])))
        elif y == 22:
            pb_row22.append((pid, float(p["x"]), float(p["y"])))
        elif y == 229:
            pb_row229.append((pid, float(p["x"]), float(p["y"])))
        elif y == 235:
            pb_row235.append((pid, float(p["x"]), float(p["y"])))

    for gid, gx, gy in grey_h:
        ry = int(round(gy))
        if 17 <= ry <= 21:
            for pid in nearest_by_x_ids(pb_row16, gx, cap=2):
                add_edge(next_map, gid, pid); add_edge(next_map, pid, gid)
            for pid in nearest_by_x_ids(pb_row22, gx, cap=2):
                add_edge(next_map, gid, pid); add_edge(next_map, pid, gid)
        else:
            for pid in nearest_by_x_ids(pb_row229, gx, cap=2):
                add_edge(next_map, gid, pid); add_edge(next_map, pid, gid)
            for pid in nearest_by_x_ids(pb_row235, gx, cap=2):
                add_edge(next_map, gid, pid); add_edge(next_map, pid, gid)

    # --- Force row22/row229 blue (E..AIL) to link to grey_h above/below by nearest X ---
    grey_band_17_21 = [(gid, gx, gy) for (gid, gx, gy) in grey_h if 17 <= int(round(gy)) <= 21]
    grey_band_230_234 = [(gid, gx, gy) for (gid, gx, gy) in grey_h if 230 <= int(round(gy)) <= 234]

    def nearest_grey_by_x_in_band(bx, band, cap=2):
        if not band:
            return []
        best = None
        best_ids = []
        for gid, gx, gy in band:
            d = abs(gx - bx)
            if best is None or d < best - EPS:
                best = d
                best_ids = [gid]
            elif abs(d - best) <= EPS:
                best_ids.append(gid)
        out = []
        for gid in best_ids:
            if gid not in out:
                out.append(gid)
        return out[:cap]

    for pid, p in points.items():
        if p["region"] != "blue":
            continue
        if p.get("meta", {}).get("kind") != "cell":
            continue
        x = int(round(float(p["x"])))
        y = int(round(float(p["y"])))
        if x < x_E or x > x_AIL:
            continue
        bx = float(p["x"])
        if y == 22:
            for gid in nearest_grey_by_x_in_band(bx, grey_band_17_21, cap=2):
                add_edge(next_map, pid, gid); add_edge(next_map, gid, pid)
        elif y == 229:
            for gid in nearest_grey_by_x_in_band(bx, grey_band_230_234, cap=2):
                add_edge(next_map, pid, gid); add_edge(next_map, gid, pid)

    # --- Step 2: AIV BLUE points with NO grey neighbor -> nearest right grey(s) (1 or 2 ties), bidirectional ---
    def has_any_grey_neighbor(pid: str) -> bool:
        outs = next_map.get(pid, set())
        return bool(outs) and any(points[t]["region"] == "grey" for t in outs)

    for pid, p in points.items():
        if p["region"] != "blue":
            continue
        if int(round(p["x"])) != x_AIV:
            continue  # ONLY AIV blue
        if has_any_grey_neighbor(pid):
            continue

        by_int = int(round(float(p["y"])))
        if 102 <= by_int <= 149:
            continue  # middle AIV blue does NOT connect to grey

        bx, by = float(p["x"]), float(p["y"])
        candidates = [(gid, gx, gy) for (gid, gx, gy) in all_grey if gx > bx + EPS]
        if not candidates:
            continue

        best = None
        best_ids = []
        for gid, gx, gy in candidates:
            d = (gx - bx) ** 2 + (gy - by) ** 2
            if best is None or d < best - EPS:
                best = d
                best_ids = [gid]
            elif abs(d - best) <= EPS:
                best_ids.append(gid)

        uniq = []
        for gid in best_ids:
            if gid not in uniq:
                uniq.append(gid)

        for gid in uniq[:2]:
            add_edge(next_map, pid, gid)
            add_edge(next_map, gid, pid)

    # -------- End right-side column AIV-AJB linking --------

    # ---------------- In/Out (parking / interaction) flag ----------------
    # Points that can serve as in/out interface nodes.
    # Stored as boolean attribute: p["inout"] = True/False.
    inout_rows_purple = {13, 14, 15, 16, 235, 236, 237, 238}
    inout_rows_orange = {31, 54, 57, 80, 83, 106, 109, 142, 145, 168, 171, 194, 197, 220}
    inout_rows_green  = {120, 131}

    x_AJB_inout = column_index_from_string("AJB") - 1
    x_AJE_inout = column_index_from_string("AJE") - 1

    for pid, p in points.items():
        x = int(round(float(p["x"])))
        y = int(round(float(p["y"])))
        reg = p.get("region", "")
        flag = False

        # All grey points (both horizontal and vertical)
        if reg == "grey":
            flag = True

        # Purple rows + AJB-AJE vertical purple strip
        elif reg == "purple":
            if y in inout_rows_purple:
                flag = True
            if x_AJB_inout <= x <= x_AJE_inout:
                flag = True

        # Orange rows
        elif reg == "orange" and y in inout_rows_orange:
            flag = True

        # Green rows
        elif reg == "green" and y in inout_rows_green:
            flag = True

        p["inout"] = bool(flag)



    # Attach next lists

    for pid, p in points.items():
        p["next"] = sorted(next_map.get(pid, set()))

    meta = {
        "excel": "F3 Traffic Config File.xlsx",
        "sheet": SHEET_NAME,
        "steps": {
            "purple_horizontal": step_purple,
            "grey": step_grey,
            "orange": step_orange,
            "green": step_green,
            "vertical_purple": step_v_purple,
            "vertical_grey": step_v_grey,
        },
        "notes": [
            "Connectivity edges are directed.",
            "Lane-change edges are created only when exact x (or y) matches exist.",
            "Vertical edges are connected only within each vertical range to avoid linking the two separated segments.",
            "Plots draw arrows only when BOTH endpoints lie inside the requested box.",
        ],
    }
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump({"meta": meta, "points": list(points.values())}, f, ensure_ascii=False, indent=2)

    # -------- Single plot: draw arrows for 4 regions on ONE figure --------
    color_map = {"blue": "blue", "green": "green", "grey": "grey", "orange": "orange", "purple": "purple"}
    by_region = defaultdict(list)
    for p in points.values():
        by_region[p["region"]].append((p["x"], p["y"], p.get("inout", False)))

    all_edges = [(p["id"], q) for p in points.values() for q in p.get("next", [])]

    # 4 regions
    plot_regions = [
        ("R1", "A12", "CO60"),
        ("R2", "A105", "CO146"),
        ("R3", "A193", "CO239"),
        ("R4", "AIE12", "AJE239"),
    ]

    # Pre-compute boxes
    boxes = []
    for name, tl, br in plot_regions:
        x0, y0 = cell_to_xy(tl)
        x1, y1 = cell_to_xy(br)
        x_min, x_max = (x0, x1) if x0 <= x1 else (x1, x0)
        y_min, y_max = (y0, y1) if y0 <= y1 else (y1, y0)
        boxes.append((name, tl, br, x_min, y_min, x_max, y_max))

    ax = plt.gca()

    # scatter all points
    for region, xy in sorted(by_region.items()):
        xs = [t[0] for t in xy]
        ys = [t[1] for t in xy]
        sizes = [3 * (9 if t[2] else 1) for t in xy]
        ax.scatter(xs, ys, s=sizes, color=color_map[region], label=f"{region} ({len(xy)})", alpha=0.85, zorder=2)

    # draw arrows for ALL regions on the same plot (union of arrows from those boxes)
    drawn_by_region = {}
    
    drawn_edges = set()
    
    for name, tl, br, x_min, y_min, x_max, y_max in boxes:
        drawn = 0
        for a, b in all_edges:
            pa = points[a]
            pb = points[b]
            if in_box(pa, x_min, y_min, x_max, y_max) or in_box(pb, x_min, y_min, x_max, y_max):
                key = (a, b)
                if key not in drawn_edges:
                    arrow = FancyArrowPatch(
                        (pa["x"], pa["y"]),
                        (pb["x"], pb["y"]),
                        arrowstyle="->",
                        mutation_scale=8,
                        linewidth=0.8,
                        alpha=0.75,
                        color="black",
                        zorder=5,
                    )
                    ax.add_patch(arrow)
                    drawn_edges.add(key)
                    drawn += 1

        drawn_by_region[name] = drawn

    ax.invert_yaxis()
    ax.set_xlabel("X (col index, A=0)")
    ax.set_ylabel("Y (row index)")
    ax.set_title("All control points + arrows within specified regions")
    #ax.legend(loc="upper right", fontsize=8, markerscale=2, frameon=True)
    legend_handles = [
        Line2D([0], [0], marker='o', color='w',
               label='blue',   markerfacecolor=color_map['blue'],   markersize=6),
        Line2D([0], [0], marker='o', color='w',
               label='green',  markerfacecolor=color_map['green'],  markersize=6),
        Line2D([0], [0], marker='o', color='w',
               label='grey',   markerfacecolor=color_map['grey'],   markersize=6),
        Line2D([0], [0], marker='o', color='w',
               label='orange', markerfacecolor=color_map['orange'], markersize=6),
        Line2D([0], [0], marker='o', color='w',
               label='purple', markerfacecolor=color_map['purple'], markersize=6),
    ]
    
    ax.legend(handles=legend_handles, loc="upper right")

    for k, v in drawn_by_region.items():
        print(f"  {k}: arrows drawn = {v:,}")
    
    print(f"Generated points: {len(points):,}")
    print(f"Generated directed edges: {len(all_edges):,}")
    print(f"JSON: {OUT_JSON}")
    
    
if __name__ == "__main__":
    main()
    
    
    